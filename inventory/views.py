import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from django.http import HttpResponse  
import re
from datetime import timedelta
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from django.contrib.auth import get_user_model
from django.contrib import messages  
from django.core.paginator import Paginator
from django.db.models import (
    BooleanField,
    Case,
    DecimalField,
    ExpressionWrapper,
    F,
    IntegerField,
    Q,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, redirect, render

from .forms import IngredientForm, InventoryLogForm
from .models import Ingredient, InventoryLog


User = get_user_model()

DECIMAL_ZERO = Value(
    Decimal("0.00"),
    output_field=DecimalField(max_digits=12, decimal_places=2),
)


def _to_decimal(value, default="0"):
    if value is None or value == "":
        return Decimal(default)

    try:
        cleaned = str(value).replace(",", "").strip()
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _clean_header(value):
    if value is None:
        return ""

    return re.sub(r"\s*\(.*\)", "", str(value)).strip()


def get_ingredient_stock_queryset():
    """재고 로그를 합산하여 실시간 현재고와 부족 상태를 계산한다."""
    return (
        Ingredient.objects.all()
        .annotate(
            in_qty=Coalesce(
                Sum(
                    Case(
                        When(inventorylog__log_type="입고", then=F("inventorylog__quantity")),
                        default=DECIMAL_ZERO,
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                DECIMAL_ZERO,
            ),
            out_qty=Coalesce(
                Sum(
                    Case(
                        When(inventorylog__log_type="출고", then=F("inventorylog__quantity")),
                        default=DECIMAL_ZERO,
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                DECIMAL_ZERO,
            ),
            waste_qty=Coalesce(
                Sum(
                    Case(
                        When(inventorylog__log_type="폐기", then=F("inventorylog__quantity")),
                        default=DECIMAL_ZERO,
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                DECIMAL_ZERO,
            ),
            adj_qty=Coalesce(
                Sum(
                    Case(
                        When(inventorylog__log_type="조정", then=F("inventorylog__quantity")),
                        default=DECIMAL_ZERO,
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                DECIMAL_ZERO,
            ),
        )
        .annotate(
            calc_stock=ExpressionWrapper(
                F("in_qty") - F("out_qty") - F("waste_qty") + F("adj_qty"),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .annotate(
            is_low_stock=Case(
                When(calc_stock__lte=F("safe_stock_level"), then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            )
        )
    )


def get_current_stock(ingredient_id):
    ingredient = get_ingredient_stock_queryset().get(pk=ingredient_id)
    return ingredient.calc_stock


def ingredient_list(request):
    """식자재 목록 보기."""
    selected_category = request.GET.get("category", "").strip()
    selected_supplier = request.GET.get("supplier", "").strip()
    search_query = request.GET.get("search", "").strip()
    sort_by = request.GET.get("sort", "status_priority")

    ingredients = get_ingredient_stock_queryset()

    if selected_category:
        ingredients = ingredients.filter(category=selected_category)

    if selected_supplier:
        ingredients = ingredients.filter(supplier__username=selected_supplier)

    if search_query:
        ingredients = ingredients.filter(name__icontains=search_query)

    low_stock_ids = ingredients.filter(calc_stock__lt=F("safe_stock_level")).values_list(
        "id",
        flat=True,
    )
    all_low_stock_ids = ",".join(map(str, low_stock_ids))

    ingredients = ingredients.annotate(
        status_priority=Case(
            When(calc_stock__lt=F("safe_stock_level"), then=Value(0)),
            default=Value(1),
            output_field=IntegerField(),
        )
    )

    if sort_by == "supplier":
        ingredients = ingredients.order_by("status_priority", "supplier__username", "id")
    elif sort_by == "name":
        ingredients = ingredients.order_by("status_priority", "name")
    else:
        ingredients = ingredients.order_by("status_priority", "-id")

    categories = (
        Ingredient.objects.exclude(category__isnull=True)
        .exclude(category__exact="")
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )
    suppliers = User.objects.filter(supplied_ingredients__isnull=False).distinct()

    paginator = Paginator(ingredients, 10)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    context = {
        "page_obj": page_obj,
        "categories": categories,
        "suppliers": suppliers,
        "selected_category": selected_category,
        "selected_supplier": selected_supplier,
        "search_query": search_query,
        "sort_by": sort_by,
        "all_low_stock_ids": all_low_stock_ids,
    }

    return render(request, "inventory/ingredient_list.html", context)


def ingredient_create(request):
    if request.method == "POST":
        form = IngredientForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "식재료가 성공적으로 등록되었습니다.")
            return redirect("inventory:ingredient_list")
    else:
        form = IngredientForm()

    return render(request, "inventory/ingredient_form.html", {"form": form})


def ingredient_upload(request):
    """식자재 기준정보를 엑셀 파일로 일괄 등록/수정한다."""
    if request.method == "GET":
        return render(request, "inventory/ingredient_upload.html")

    excel_file = request.FILES.get("file")

    if not excel_file:
        messages.error(request, "업로드할 엑셀(.xlsx) 파일을 선택해 주세요.")
        return redirect("inventory:ingredient_upload")

    if not excel_file.name.lower().endswith(".xlsx"):
        messages.error(request, "현재는 .xlsx 파일만 지원합니다.")
        return redirect("inventory:ingredient_upload")

    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook.active

        headers = [_clean_header(cell.value) for cell in sheet[1]]

        col_map = {
            "식재료명": "name",
            "발주처": "supplier",
            "규격": "spec",
            "단위": "unit",
            "단가": "unit_price",
            "안전재고": "safe_stock_level",
            "대분류": "category",
            "연간예상소요량": "yearly_demand",
            "금액": "total_amount",
            "식품 설명": "description",
            "설명": "description",
        }

        positions = {}
        for excel_name, field_name in col_map.items():
            if excel_name in headers and field_name not in positions:
                positions[field_name] = headers.index(excel_name)

        if "name" not in positions:
            messages.error(request, "엑셀 파일에 '식재료명' 컬럼이 포함되어야 합니다.")
            return redirect("inventory:ingredient_upload")

        success_count = 0
        skipped_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            def get_value(field_name):
                index = positions.get(field_name)
                return row[index] if index is not None else None

            name = get_value("name")

            if not name:
                skipped_count += 1
                continue

            supplier_value = get_value("supplier")
            supplier_obj = None

            if supplier_value:
                supplier_obj = User.objects.filter(
                    Q(username=str(supplier_value).strip())
                    | Q(name=str(supplier_value).strip())
                ).first()

            defaults = {
                "supplier": supplier_obj,
                "spec": get_value("spec") or "",
                "unit": get_value("unit") or "EA",
                "unit_price": int(_to_decimal(get_value("unit_price"))),
                "safe_stock_level": _to_decimal(get_value("safe_stock_level")),
                "category": get_value("category") or "",
                "yearly_demand": int(_to_decimal(get_value("yearly_demand"))),
                "total_amount": int(_to_decimal(get_value("total_amount"))),
            }

            description = get_value("description")
            if description is not None:
                defaults["description"] = str(description)

            Ingredient.objects.update_or_create(
                name=str(name).strip(),
                defaults=defaults,
            )
            success_count += 1

        messages.success(
            request,
            f"식자재 엑셀 업로드 완료: {success_count}건 처리, {skipped_count}건 건너뜀.",
        )

    except Exception as exc:
        messages.error(request, f"엑셀 업로드 중 오류가 발생했습니다: {exc}")
        return redirect("inventory:ingredient_upload")

    return redirect("inventory:ingredient_list")


def inventory_log_create(request):
    all_ingredients = get_ingredient_stock_queryset()
    recent_logs = (
        InventoryLog.objects.filter(log_type__in=["출고", "폐기"])
        .select_related("ingredient")
        .order_by("-transaction_date", "-id")[:15]
    )

    if request.method == "POST":
        form = InventoryLogForm(request.POST)

        if form.is_valid():
            ingredient = form.cleaned_data["ingredient"]
            quantity = form.cleaned_data["quantity"]
            log_type = form.cleaned_data["log_type"]
            current_stock = get_current_stock(ingredient.id)

            if log_type in ["출고", "폐기"] and quantity > current_stock:
                form.add_error("quantity", f"현재고({current_stock})가 부족합니다.")
            else:
                form.save()
                messages.success(request, f"[{ingredient.name}] {log_type} 기록 완료.")
                return redirect("inventory:inventory_log_create")
    else:
        form = InventoryLogForm()

    return render(
        request,
        "inventory/inventory_log_form.html",
        {
            "form": form,
            "recent_logs": recent_logs,
            "ingredients": all_ingredients,
        },
    )


def ingredient_stock_adjust(request, pk=None):
    """수동 재고 보정."""
    if request.method != "POST":
        return redirect("inventory:ingredient_list")

    ingredient = get_object_or_404(Ingredient, pk=pk)

    try:
        new_stock = Decimal(request.POST.get("new_stock", 0))
        current_val = get_current_stock(ingredient.id)
        diff = new_stock - current_val

        if diff != 0:
            InventoryLog.objects.create(
                ingredient=ingredient,
                log_type="조정",
                quantity=diff,
                description="수동 보정",
            )

        messages.success(request, f"{ingredient.name} 재고가 {new_stock}으로 수동 보정되었습니다.")

    except Exception as exc:
        messages.error(request, f"오류가 발생했습니다: {exc}")

    return redirect("inventory:ingredient_list")

# --- 수불 현황 보고서 및 엑셀 출력 ---

def weekly_inventory_report(request):
    """주간수불명세서 조회 뷰 (월~일 집계)"""
    target_date_str = request.GET.get('date', timezone.now().date().isoformat())
    try:
        requested_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    except ValueError:
        requested_date = timezone.now().date()

    # 해당 주의 월요일 계산
    monday = requested_date - timedelta(days=requested_date.weekday())
    sunday = monday + timedelta(days=6)
    
    # 시간대 고려한 범위 설정 (월요일 00:00:00 ~ 일요일 23:59:59)
    start_dt = timezone.make_aware(datetime.combine(monday, datetime.min.time()))
    end_dt = timezone.make_aware(datetime.combine(sunday, datetime.max.time()))

    ingredients = Ingredient.objects.all().order_by('category', 'name')
    report_data = []

    for item in ingredients:
        # 기초 재고 (월요일 시작 전)
        prev_stock = InventoryLog.objects.filter(ingredient=item, transaction_date__lt=start_dt).aggregate(
            total=Sum(Case(
                When(log_type='입고', then=F('quantity')),
                When(log_type='출고', then=-F('quantity')),
                When(log_type='폐기', then=-F('quantity')),
                When(log_type='조정', then=F('quantity')),
                default=Value(0),
                output_field=DecimalField()
            ))
        )['total'] or Decimal('0.00')

        # 주간 수불 내역 (%)
        week_logs = InventoryLog.objects.filter(ingredient=item, transaction_date__range=(start_dt, end_dt))
        in_qty = week_logs.filter(log_type='입고').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')
        out_qty = (week_logs.filter(log_type='출고').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')) + \
                  (week_logs.filter(log_type='폐기').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00'))
        adj_qty = week_logs.filter(log_type='조정').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')

        curr_stock = prev_stock + in_qty - out_qty + adj_qty

        if prev_stock != 0 or in_qty != 0 or out_qty != 0 or adj_qty != 0:
            report_data.append({
                'ingredient': item,
                'prev_stock': prev_stock,
                'in_qty': in_qty,
                'out_qty': out_qty,
                'adj_qty': adj_qty,
                'curr_stock': curr_stock,
            })

    context = {
        'monday': monday,
        'sunday': sunday,
        'target_date': requested_date,
        'report_data': report_data,
        'prev_week_date': (monday - timedelta(days=7)).isoformat(),
        'next_week_date': (monday + timedelta(days=7)).isoformat(),
    }
    return render(request, 'docs/weekly_inventory_report.html', context)

def monthly_inventory_report(request):
    """월수불명세서 조회 뷰"""
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    start_dt = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_dt = timezone.make_aware(datetime.combine(end_date, datetime.min.time()))

    ingredients = Ingredient.objects.all().order_by('category', 'name')
    report_data = []

    for item in ingredients:
        # 기초 재고 (월초 이전)
        prev_stock = InventoryLog.objects.filter(ingredient=item, transaction_date__lt=start_dt).aggregate(
            total=Sum(Case(
                When(log_type='입고', then=F('quantity')),
                When(log_type='출고', then=-F('quantity')),
                When(log_type='폐기', then=-F('quantity')),
                When(log_type='조정', then=F('quantity')),
                default=Value(0),
                output_field=DecimalField()
            ))
        )['total'] or Decimal('0.00')

        # 당월 수불 내역
        month_logs = InventoryLog.objects.filter(ingredient=item, transaction_date__range=(start_dt, end_dt))
        in_qty = month_logs.filter(log_type='입고').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')
        out_qty = (month_logs.filter(log_type='출고').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')) + \
                  (month_logs.filter(log_type='폐기').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00'))
        adj_qty = month_logs.filter(log_type='조정').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')

        curr_stock = prev_stock + in_qty - out_qty + adj_qty

        if prev_stock != 0 or in_qty != 0 or out_qty != 0 or adj_qty != 0:
            report_data.append({
                'ingredient': item,
                'prev_stock': prev_stock,
                'in_qty': in_qty,
                'out_qty': out_qty,
                'adj_qty': adj_qty,
                'curr_stock': curr_stock,
            })

    context = {
        'year': year,
        'month': month,
        'report_data': report_data,
    }
    return render(request, 'docs/monthly_inventory_report.html', context)

def export_inventory_excel(request):
    """수불 내역 엑셀 다운로드 (주간/월간 지원)"""
    report_type = request.GET.get('type', 'weekly')
    target_date_str = request.GET.get('date', timezone.now().date().isoformat())
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수불명세서"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["No", "분류", "식재료명", "규격", "단위", "기초재고", "입고합계", "출고합계", "보정액", "기말재고", "비고"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # 범위 설정
    if report_type == 'monthly':
        year = int(request.GET.get('year', timezone.now().year))
        month = int(request.GET.get('month', timezone.now().month))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        filename = f"Inventory_Report_Monthly_{year}_{month}.xlsx"
    else: # weekly
        requested_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        monday = requested_date - timedelta(days=requested_date.weekday())
        sunday = monday + timedelta(days=6)
        start_date = monday
        end_date = sunday + timedelta(days=1)
        filename = f"Inventory_Report_Weekly_{monday.isoformat()}.xlsx"

    start_dt = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    if report_type == 'monthly':
        end_dt = timezone.make_aware(datetime.combine(end_date, datetime.min.time()))
    else:
        end_dt = timezone.make_aware(datetime.combine(sunday, datetime.max.time()))

    ingredients = Ingredient.objects.all().order_by('category', 'name')
    row_num = 2
    for idx, item in enumerate(ingredients, 1):
        # 기초 재고
        prev_stock = InventoryLog.objects.filter(ingredient=item, transaction_date__lt=start_dt).aggregate(
            total=Sum(Case(
                When(log_type='입고', then=F('quantity')),
                When(log_type='출고', then=-F('quantity')),
                When(log_type='폐기', then=-F('quantity')),
                When(log_type='조정', then=F('quantity')),
                default=Value(0),
                output_field=DecimalField()
            ))
        )['total'] or Decimal('0.00')

        # 범위 내 수불
        logs = InventoryLog.objects.filter(ingredient=item, transaction_date__range=(start_dt, end_dt))
        in_qty = logs.filter(log_type='입고').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')
        out_qty = (logs.filter(log_type='출고').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')) + \
                  (logs.filter(log_type='폐기').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00'))
        adj_qty = logs.filter(log_type='조정').aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0.00')
        curr_stock = prev_stock + in_qty - out_qty + adj_qty

        if prev_stock != 0 or in_qty != 0 or out_qty != 0 or adj_qty != 0:
            ws.cell(row=row_num, column=1, value=row_num-1).border = thin_border
            ws.cell(row=row_num, column=2, value=item.category).border = thin_border
            ws.cell(row=row_num, column=3, value=item.name).border = thin_border
            ws.cell(row=row_num, column=4, value=item.spec).border = thin_border
            ws.cell(row=row_num, column=5, value=item.unit).border = thin_border
            ws.cell(row=row_num, column=6, value=prev_stock).border = thin_border
            ws.cell(row=row_num, column=7, value=in_qty).border = thin_border
            ws.cell(row=row_num, column=8, value=out_qty).border = thin_border
            ws.cell(row=row_num, column=9, value=adj_qty).border = thin_border
            ws.cell(row=row_num, column=10, value=curr_stock).border = thin_border
            ws.cell(row=row_num, column=11, value="").border = thin_border
            row_num += 1

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response