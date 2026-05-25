import json
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.http import require_GET, require_POST

from inventory.models import Ingredient

from .models import Menu, Recipe


HEADER_ALIASES = {
    "menu": ["메뉴명", "메뉴", "음식명", "요리명"],
    "category": ["카테고리", "분류", "메뉴분류", "메뉴 분류"],
    "ingredient": ["식재료명", "식재료", "재료명", "재료"],
    "amount": ["사용량", "소요량", "1인분소요량", "1인분 사용량", "required_amount"],
}


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _compact_name(value):
    return _normalize_text(value).replace(" ", "")


def _to_decimal(value):
    if value is None or value == "":
        return None

    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _find_header(sheet):
    """상위 20행에서 메뉴명/식재료명/사용량 헤더를 찾아 컬럼 위치를 반환한다."""
    max_scan_row = min(sheet.max_row, 20)

    for row_index in range(1, max_scan_row + 1):
        row_values = [
            _normalize_text(sheet.cell(row=row_index, column=col).value)
            for col in range(1, sheet.max_column + 1)
        ]
        compact_values = [value.replace(" ", "") for value in row_values]

        col_map = {}

        for key, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                compact_alias = alias.replace(" ", "")
                if compact_alias in compact_values:
                    col_map[key] = compact_values.index(compact_alias)
                    break

        if "menu" in col_map and "ingredient" in col_map and "amount" in col_map:
            return row_index, col_map

    # 재석 브랜치의 fallback과 호환: 4행을 헤더로 보고 메뉴/카테고리/식재료/사용량 순서 가정
    return 4, {"menu": 0, "category": 1, "ingredient": 2, "amount": 3}


def _find_ingredient_by_name(name):
    compact = _compact_name(name)

    if not compact:
        return None

    for ingredient in Ingredient.objects.all():
        if _compact_name(ingredient.name) == compact:
            return ingredient

    return Ingredient.objects.filter(name__icontains=_normalize_text(name)).first()


def _get_or_create_ingredient(name, created_names):
    ingredient = _find_ingredient_by_name(name)

    if ingredient:
        return ingredient

    ingredient = Ingredient.objects.create(
        name=_normalize_text(name),
        category="기타",
        unit="kg",
        unit_price=0,
        safe_stock_level=0,
    )
    created_names.append(ingredient.name)

    return ingredient


@require_POST
@transaction.atomic
def recipe_upload_api(request):
    """엑셀 파일로 메뉴-식재료-사용량 레시피를 일괄 등록한다.

    기대 컬럼:
    - 메뉴명
    - 카테고리
    - 식재료명
    - 사용량

    컬럼 순서가 바뀌어도 헤더명으로 매칭한다.
    식재료가 없으면 기타/kg/안전재고0 기준으로 자동 생성한다.
    """
    excel_file = request.FILES.get("excel_file")

    if not excel_file:
        return JsonResponse(
            {"ok": False, "message": "업로드할 엑셀 파일을 선택해 주세요."},
            status=400,
        )

    if not excel_file.name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return JsonResponse(
            {"ok": False, "message": "현재는 .xlsx 계열 엑셀 파일만 지원합니다."},
            status=400,
        )

    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        return JsonResponse(
            {"ok": False, "message": f"엑셀 파일을 읽을 수 없습니다: {exc}"},
            status=400,
        )

    header_row, col_map = _find_header(sheet)
    start_row = header_row + 1

    processed_count = 0
    skipped_rows = []
    created_menus = []
    created_ingredients = []

    for row_num in range(start_row, sheet.max_row + 1):
        row_values = [
            sheet.cell(row=row_num, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        def value_for(key):
            index = col_map.get(key)
            if index is None or index >= len(row_values):
                return None
            return row_values[index]

        menu_name = _normalize_text(value_for("menu"))
        category = _normalize_text(value_for("category")) or "기타"
        ingredient_name = _normalize_text(value_for("ingredient"))
        amount = _to_decimal(value_for("amount"))

        if not menu_name and not ingredient_name:
            continue

        if not menu_name or not ingredient_name or amount is None:
            skipped_rows.append(
                {
                    "row": row_num,
                    "menu_name": menu_name,
                    "ingredient_name": ingredient_name,
                    "amount": _normalize_text(value_for("amount")),
                    "reason": "메뉴명/식재료명/사용량 누락 또는 사용량 숫자 변환 실패",
                }
            )
            continue

        menu, menu_created = Menu.objects.get_or_create(
            name=menu_name,
            defaults={"category": category},
        )

        if not menu_created and category and not menu.category:
            menu.category = category
            menu.save(update_fields=["category"])

        if menu_created:
            created_menus.append(menu.name)

        ingredient = _get_or_create_ingredient(ingredient_name, created_ingredients)

        Recipe.objects.update_or_create(
            menu=menu,
            ingredient=ingredient,
            defaults={"required_amount": amount},
        )

        processed_count += 1

    return JsonResponse(
        {
            "ok": True,
            "message": f"레시피 엑셀 업로드 완료: {processed_count}건 반영",
            "processed_count": processed_count,
            "skipped_count": len(skipped_rows),
            "skipped_rows": skipped_rows[:30],
            "created_menus": created_menus,
            "created_ingredients": created_ingredients,
        }
    )


@require_GET
def search_ingredients_api(request):
    query = _normalize_text(request.GET.get("q", ""))

    ingredients = Ingredient.objects.all().order_by("name")

    if query:
        compact_query = _compact_name(query)
        matched_ids = []

        for ingredient in ingredients:
            if compact_query in _compact_name(ingredient.name):
                matched_ids.append(ingredient.id)

        ingredients = Ingredient.objects.filter(id__in=matched_ids).order_by("name")

    results = [
        {
            "id": ingredient.id,
            "name": ingredient.name,
            "unit": ingredient.unit,
            "category": ingredient.category or "",
        }
        for ingredient in ingredients[:20]
    ]

    return JsonResponse({"results": results})


@require_GET
def get_menu_recipes_api(request, menu_id):
    try:
        menu = Menu.objects.prefetch_related("recipes__ingredient").get(id=menu_id)
    except Menu.DoesNotExist:
        return JsonResponse({"ok": False, "message": "메뉴를 찾을 수 없습니다."}, status=404)

    recipes = [
        {
            "ingredient_id": recipe.ingredient_id,
            "ingredient_name": recipe.ingredient.name,
            "required_amount": str(recipe.required_amount),
            "unit": recipe.ingredient.unit,
        }
        for recipe in menu.recipes.all().order_by("ingredient__name")
    ]

    return JsonResponse(
        {
            "ok": True,
            "menu": {
                "id": menu.id,
                "name": menu.name,
                "category": menu.category or "",
            },
            "recipes": recipes,
        }
    )
