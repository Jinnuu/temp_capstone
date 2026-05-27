import calendar
import json
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.core.management import call_command
from django.db import transaction
from django.db.models import Case, IntegerField, Value, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from inventory.models import Ingredient
from .models import DietMenu, DietPlan, Menu, Recipe


CATEGORY_ORDER = ["밥", "국", "주반찬", "부반찬", "김치", "간식"]

SORT_ORDER = Case(
    *[When(menu__category=cat, then=Value(pos)) for pos, cat in enumerate(CATEGORY_ORDER)],
    default=Value(len(CATEGORY_ORDER)),
    output_field=IntegerField(),
)


def to_decimal(value):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def meal_home(request):
    return render(request, "meals/meal_home.html")


def mealplan_create(request):
    menus = Menu.objects.all().order_by("name")
    selected_date = request.GET.get("date") or request.POST.get("meal_date")

    existing_meals = {
        "breakfast": [],
        "lunch": [],
        "dinner": [],
    }

    if selected_date:
        plans = (
            DietPlan.objects.filter(target_date=selected_date)
            .prefetch_related("diet_menus__menu")
            .order_by("meal_type")
        )

        for plan in plans:
            menu_ids = [item.menu_id for item in plan.diet_menus.all()]

            if plan.meal_type == "조식":
                existing_meals["breakfast"] = menu_ids
            elif plan.meal_type == "중식":
                existing_meals["lunch"] = menu_ids
            elif plan.meal_type == "석식":
                existing_meals["dinner"] = menu_ids

    if request.method == "POST":
        meal_date = request.POST.get("meal_date")
        breakfast_menu_ids = request.POST.getlist("breakfast_menu")
        lunch_menu_ids = request.POST.getlist("lunch_menu")
        dinner_menu_ids = request.POST.getlist("dinner_menu")

        if not meal_date:
            messages.error(request, "날짜를 선택해주세요.")
            return redirect("meals:mealplan_create")

        DietPlan.objects.filter(target_date=meal_date).delete()

        created_count = 0
        meal_type_map = {
            "조식": breakfast_menu_ids,
            "중식": lunch_menu_ids,
            "석식": dinner_menu_ids,
        }

        for meal_type, menu_ids in meal_type_map.items():
            valid_menu_ids = [menu_id for menu_id in menu_ids if menu_id]

            if valid_menu_ids:
                diet_plan = DietPlan.objects.create(
                    target_date=meal_date,
                    meal_type=meal_type,
                )

                for menu_id in valid_menu_ids:
                    DietMenu.objects.create(
                        diet_plan=diet_plan,
                        menu_id=menu_id,
                    )

                created_count += 1

        if created_count == 0:
            messages.warning(request, "선택된 메뉴가 없어 빈 식단으로 저장되지 않았습니다.")
        else:
            messages.success(request, f"{meal_date} 식단이 저장되었습니다.")

        return redirect(f"{reverse('meals:mealplan_create')}?date={meal_date}")

    context = {
        "menus": menus,
        "selected_date": selected_date,
        "existing_meals": existing_meals,
    }

    return render(request, "meals/mealplan_create.html", context)


def mealplan_list(request):
    today = date.today()
    selected_year = int(request.GET.get("year", today.year))
    selected_month = int(request.GET.get("month", today.month))

    plans = (
        DietPlan.objects.filter(
            target_date__year=selected_year,
            target_date__month=selected_month,
        )
        .prefetch_related("diet_menus__menu")
        .order_by("target_date", "meal_type")
    )

    meal_map = {}

    for plan in plans:
        day = plan.target_date.day
        meal_map.setdefault(
            day,
            {
                "breakfast": [],
                "lunch": [],
                "dinner": [],
                "breakfast_dicts": [],
                "lunch_dicts": [],
                "dinner_dicts": [],
            },
        )

        sorted_diet_menus = plan.diet_menus.all().select_related("menu").order_by(SORT_ORDER)
        menu_names = [item.menu.name for item in sorted_diet_menus]
        menu_dicts = [{"id": item.menu.id, "name": item.menu.name} for item in sorted_diet_menus]

        if plan.meal_type == "조식":
            meal_map[day]["breakfast"] = menu_names
            meal_map[day]["breakfast_dicts"] = json.dumps(menu_dicts)
        elif plan.meal_type == "중식":
            meal_map[day]["lunch"] = menu_names
            meal_map[day]["lunch_dicts"] = json.dumps(menu_dicts)
        elif plan.meal_type == "석식":
            meal_map[day]["dinner"] = menu_names
            meal_map[day]["dinner_dicts"] = json.dumps(menu_dicts)

    calendar_weeks = []
    month_calendar = calendar.Calendar(firstweekday=calendar.SUNDAY)

    for week in month_calendar.monthdayscalendar(selected_year, selected_month):
        week_cells = []

        for day_num in week:
            if day_num == 0:
                week_cells.append(None)
                continue

            week_cells.append(
                {
                    "day": day_num,
                    "date": date(selected_year, selected_month, day_num),
                    "is_today": (
                        today.year == selected_year
                        and today.month == selected_month
                        and today.day == day_num
                    ),
                    "meals": meal_map.get(
                        day_num,
                        {
                            "breakfast": [],
                            "lunch": [],
                            "dinner": [],
                            "breakfast_dicts": [],
                            "lunch_dicts": [],
                            "dinner_dicts": [],
                        },
                    ),
                }
            )

        calendar_weeks.append(week_cells)

    context = {
        "selected_year": selected_year,
        "selected_month": selected_month,
        "months": range(1, 13),
        "calendar_weeks": calendar_weeks,
        "today_year": today.year,
        "today_month": today.month,
    }

    return render(request, "meals/mealplan_list.html", context)


def save_recipe_rows(menu, ingredient_ids, amounts):
    valid_count = 0
    skipped_count = 0
    seen_ingredient_ids = set()

    for ingredient_id, amount in zip(ingredient_ids, amounts):
        if not ingredient_id or not amount:
            continue

        amount_decimal = to_decimal(amount)

        if amount_decimal is None or amount_decimal <= 0:
            skipped_count += 1
            continue

        try:
            ingredient = Ingredient.objects.get(id=ingredient_id)
        except Ingredient.DoesNotExist:
            skipped_count += 1
            continue

        if ingredient.id in seen_ingredient_ids:
            skipped_count += 1
            continue

        seen_ingredient_ids.add(ingredient.id)

        Recipe.objects.update_or_create(
            menu=menu,
            ingredient=ingredient,
            defaults={"required_amount": amount_decimal},
        )

        valid_count += 1

    return valid_count, skipped_count


def recipe_create(request):
    menus = Menu.objects.all().order_by("name")
    ingredients = Ingredient.objects.all().order_by("name")

    if request.method == "POST":
        menu_id = request.POST.get("menu_id")
        ingredient_ids = request.POST.getlist("ingredient_id[]")
        required_amounts = request.POST.getlist("required_amount[]")

        if not menu_id:
            messages.error(request, "수정할 메뉴를 선택해주세요.")
            return redirect("meals:recipe_create")

        menu = get_object_or_404(Menu, id=menu_id)

        with transaction.atomic():
            Recipe.objects.filter(menu=menu).delete()
            valid_count, skipped_count = save_recipe_rows(menu, ingredient_ids, required_amounts)

        if valid_count:
            messages.success(request, f"{menu.name} 레시피 식재료 {valid_count}개가 저장되었습니다.")
        else:
            messages.warning(request, "저장된 식재료가 없습니다. 식재료 검색 후 후보를 클릭했는지 확인해주세요.")

        if skipped_count:
            messages.warning(request, f"중복/오류 입력 {skipped_count}건은 제외했습니다.")

        return redirect("meals:menu_list")

    context = {
        "menus": menus,
        "ingredients": ingredients,
    }

    return render(request, "meals/recipe_create.html", context)


def menu_create(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        category = request.POST.get("category")
        ingredient_ids = request.POST.getlist("ingredient_id[]")
        amounts = request.POST.getlist("ing_amount[]")

        if not name or not category:
            messages.error(request, "메뉴명과 카테고리를 입력해주세요.")
            return render(request, "meals/menu_create.html")

        with transaction.atomic():
            menu = Menu.objects.create(
                name=name,
                category=category,
            )
            valid_count, skipped_count = save_recipe_rows(menu, ingredient_ids, amounts)

        if valid_count:
            messages.success(request, f"{menu.name} 메뉴와 식재료 {valid_count}개가 저장되었습니다.")
        else:
            messages.warning(request, f"{menu.name} 메뉴는 저장되었지만 연결된 식재료가 없습니다. 식재료 검색 후 후보를 클릭했는지 확인해주세요.")

        if skipped_count:
            messages.warning(request, f"중복/오류 입력 {skipped_count}건은 제외했습니다.")

        return redirect("meals:menu_list")

    return render(request, "meals/menu_create.html")


def menu_list(request):
    q = request.GET.get("q", "")
    category = request.GET.get("category", "")

    menus = Menu.objects.all().prefetch_related("recipes__ingredient")

    if q:
        menus = menus.filter(name__icontains=q)

    if category:
        menus = menus.filter(category=category)

    return render(request, "meals/menu_list.html", {"menus": menus})


def menu_delete(request, menu_id):
    menu = get_object_or_404(Menu, id=menu_id)
    menu.delete()
    messages.success(request, "메뉴가 삭제되었습니다.")
    return redirect("meals:menu_list")


def menu_update(request, menu_id):
    menu = get_object_or_404(Menu.objects.prefetch_related("recipes__ingredient"), id=menu_id)

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        category = request.POST.get("category")
        ingredient_ids = request.POST.getlist("ingredient_id[]")
        amounts = request.POST.getlist("ing_amount[]")

        if not name or not category:
            messages.error(request, "메뉴명과 카테고리를 입력해주세요.")
            return render(request, "meals/menu_update.html", {"menu": menu})

        with transaction.atomic():
            menu.name = name
            menu.category = category
            menu.save()

            Recipe.objects.filter(menu=menu).delete()
            valid_count, skipped_count = save_recipe_rows(menu, ingredient_ids, amounts)

        if valid_count:
            messages.success(request, f"{menu.name} 레시피 식재료 {valid_count}개가 저장되었습니다.")
        else:
            messages.warning(request, f"{menu.name} 메뉴는 저장되었지만 연결된 식재료가 없습니다. 식재료 검색 후 후보를 클릭했는지 확인해주세요.")

        if skipped_count:
            messages.warning(request, f"중복/오류 입력 {skipped_count}건은 제외했습니다.")

        return redirect("meals:menu_list")

    return render(request, "meals/menu_update.html", {"menu": menu})


def deduct_inventory_view(request):
    if request.method == "POST":
        target_date = request.POST.get("target_date")

        if target_date:
            call_command("deduct_daily_inventory", date=target_date)
            messages.success(request, f"{target_date} 기준 식재료 일괄 사용처리가 완료되었습니다.")
        else:
            call_command("deduct_daily_inventory")
            messages.success(request, "오늘 기준 식재료 일괄 사용처리가 완료되었습니다.")

    return redirect("meals:mealplan_list")


def search_menus_api(request):
    query = request.GET.get("q", "").strip()
    menus = list(Menu.objects.all())

    if not menus:
        return JsonResponse({"results": []})

    if not query:
        results = [{"id": m.id, "name": m.name, "category": m.category or ""} for m in menus[:50]]
        return JsonResponse({"results": results})

    documents = [m.name for m in menus]
    documents.append(query)

    vectorizer = TfidfVectorizer(analyzer="char_wb", ngram_range=(1, 3))

    try:
        tfidf_matrix = vectorizer.fit_transform(documents)
    except ValueError:
        return JsonResponse({"results": []})

    cosine_sim = cosine_similarity(tfidf_matrix[-1], tfidf_matrix[:-1]).flatten()
    top_indices = cosine_sim.argsort()[-20:][::-1]

    results = []

    for idx in top_indices:
        if cosine_sim[idx] > 0.0:
            menu = menus[idx]
            results.append(
                {
                    "id": menu.id,
                    "name": menu.name,
                    "category": menu.category or "",
                    "score": float(cosine_sim[idx]),
                }
            )

    return JsonResponse({"results": results})


@csrf_exempt
def add_diet_menu_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date_str = data.get("date")
            meal_type = data.get("meal_type")
            menu_id = data.get("menu_id")
            menu_name = data.get("menu_name")

            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()

            if menu_id:
                menu = Menu.objects.get(id=menu_id)
            elif menu_name:
                menu, _ = Menu.objects.get_or_create(
                    name=menu_name,
                    defaults={"category": "기타"},
                )
            else:
                return JsonResponse({"status": "error", "message": "No menu info provided"}, status=400)

            plan, _ = DietPlan.objects.get_or_create(
                target_date=target_date,
                meal_type=meal_type,
            )

            DietMenu.objects.get_or_create(diet_plan=plan, menu=menu)

            sorted_menus = plan.diet_menus.all().select_related("menu").order_by(SORT_ORDER)
            menu_data = [{"id": dm.menu.id, "name": dm.menu.name} for dm in sorted_menus]

            return JsonResponse(
                {
                    "status": "success",
                    "menu_id": menu.id,
                    "menu_name": menu.name,
                    "menus": menu_data,
                }
            )
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error"}, status=405)


@csrf_exempt
def remove_diet_menu_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date_str = data.get("date")
            meal_type = data.get("meal_type")
            menu_id = data.get("menu_id")

            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            plan = DietPlan.objects.get(target_date=target_date, meal_type=meal_type)
            DietMenu.objects.filter(diet_plan=plan, menu_id=menu_id).delete()

            if not plan.diet_menus.exists():
                plan.delete()
                return JsonResponse({"status": "success", "menus": []})

            sorted_menus = plan.diet_menus.all().select_related("menu").order_by(SORT_ORDER)
            menu_data = [{"id": dm.menu.id, "name": dm.menu.name} for dm in sorted_menus]

            return JsonResponse({"status": "success", "menus": menu_data})
        except DietPlan.DoesNotExist:
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error"}, status=405)


def weekly_mealplan_create(request):
    start_date_str = request.GET.get("start_date")

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = date.today()
    else:
        start_date = date.today()

    start_date = start_date - timedelta(days=start_date.weekday())

    if request.method == "POST":
        post_start_date_str = request.POST.get("start_date") or start_date.strftime("%Y-%m-%d")
        messages.success(request, "완료되었습니다.")
        return redirect(f"{reverse('meals:weekly_mealplan_create')}?start_date={post_start_date_str}")

    dates = [start_date + timedelta(days=i) for i in range(7)]

    all_plans = (
        DietPlan.objects.filter(
            target_date__gte=dates[0],
            target_date__lte=dates[-1],
        )
        .prefetch_related("diet_menus__menu")
    )

    weekly_data = {}

    for d in dates:
        weekly_data[d] = {
            "조식": [],
            "중식": [],
            "석식": [],
        }

    for plan_obj in all_plans:
        if plan_obj.target_date in weekly_data:
            sorted_diet_menus = plan_obj.diet_menus.all().select_related("menu").order_by(SORT_ORDER)
            menus = [dm.menu for dm in sorted_diet_menus]
            weekly_data[plan_obj.target_date][plan_obj.meal_type] = menus

    days_context = []
    weekdays_list = ["월", "화", "수", "목", "금", "토", "일"]

    for i, d in enumerate(dates):
        days_context.append(
            {
                "date": d,
                "date_str": d.strftime("%Y-%m-%d"),
                "weekday": weekdays_list[i],
                "breakfast": weekly_data[d].get("조식", []),
                "lunch": weekly_data[d].get("중식", []),
                "dinner": weekly_data[d].get("석식", []),
                "is_next_monday": False,
            }
        )

    context = {
        "start_date": start_date,
        "start_date_str": start_date.strftime("%Y-%m-%d"),
        "prev_week": (start_date - timedelta(days=7)).strftime("%Y-%m-%d"),
        "next_week": (start_date + timedelta(days=7)).strftime("%Y-%m-%d"),
        "days_context": days_context,
    }

    return render(request, "meals/weekly_mealplan_create.html", context)


def weekly_meal_plan_document(request):
    """
    주간 식단표 문서화 기능을 위한 뷰입니다.
    ?date=YYYY-MM-DD 파라미터를 지원하며, 해당 날짜가 속한 주의 월요일을 기준으로 8일치 데이터를 조회합니다.
    """
    # 1. 파라미터 처리 (date 또는 start_date 인식)
    target_date_str = request.GET.get("date") or request.GET.get("start_date")

    # 2. 동적 날짜 계산
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    # 해당 날짜가 속한 주의 월요일(start_date) 계산
    start_date = target_date - timedelta(days=target_date.weekday())
    end_date = start_date + timedelta(days=6)  # 일요일

    # 3. 8일 데이터 바인딩 (월요일부터 다음 주 월요일까지)
    dates = [start_date + timedelta(days=i) for i in range(8)]

    # 8일간의 식단 데이터를 한 번에 가져옴
    all_plans = (
        DietPlan.objects.filter(
            target_date__gte=dates[0],
            target_date__lte=dates[7],
        )
        .prefetch_related("diet_menus__menu")
    )

    # 데이터를 날짜 및 끼니별로 매핑
    weekly_data = {d: {"조식": [], "중식": [], "석식": []} for d in dates}

    for plan_obj in all_plans:
        if plan_obj.target_date in weekly_data:
            # 카테고리 순서(SORT_ORDER)에 따라 메뉴 정렬
            sorted_diet_menus = plan_obj.diet_menus.all().select_related("menu").order_by(SORT_ORDER)
            menus = [dm.menu.name for dm in sorted_diet_menus]
            weekly_data[plan_obj.target_date][plan_obj.meal_type] = menus

    # 템플릿에 전달할 7일(월~일) 데이터 구성
    days_context = []
    weekdays_list = ["월", "화", "수", "목", "금", "토", "일"]

    for i in range(7):
        d = dates[i]
        day_info = {
            "date": d,
            "date_str": d.strftime("%Y-%m-%d"),
            "weekday": weekdays_list[i],
            "breakfast": weekly_data[d].get("조식", []),
            "lunch": weekly_data[d].get("중식", []),
            "dinner": weekly_data[d].get("석식", []),
            "remarks": [],
        }

        # 특이사항: 일요일(index 6)의 비고 칸에 다음 주 월요일 조식 데이터 매칭
        if i == 6:
            next_monday = dates[7]
            day_info["remarks"] = weekly_data[next_monday].get("조식", [])

        days_context.append(day_info)

    # 4. 네비게이션 및 상세 컨텍스트 전달
    context = {
        "start_date": start_date,
        "start_date_str": start_date.strftime("%Y-%m-%d"),
        "end_date_str": end_date.strftime("%Y-%m-%d"),
        "prev_week": (start_date - timedelta(days=7)).strftime("%Y-%m-%d"),
        "next_week": (start_date + timedelta(days=7)).strftime("%Y-%m-%d"),
        "days_context": days_context,
    }

    # 확인된 실존 템플릿 경로로 수정
    return render(request, "docs/weekly_report.html", context)


def monthly_meal_plan_document(request):
    """
    월간 식단표 문서화 기능을 위한 뷰입니다. (조회 전용)
    """
    today = date.today()
    selected_year = int(request.GET.get("year", today.year))
    selected_month = int(request.GET.get("month", today.month))

    # 해당 월의 식단 데이터 조회
    plans = (
        DietPlan.objects.filter(
            target_date__year=selected_year,
            target_date__month=selected_month,
        )
        .prefetch_related("diet_menus__menu")
        .order_by("target_date", "meal_type")
    )

    # 데이터 매핑 {day: {breakfast: [], ...}}
    meal_map = {}
    for plan in plans:
        day = plan.target_date.day
        meal_map.setdefault(day, {"breakfast": [], "lunch": [], "dinner": []})

        sorted_diet_menus = plan.diet_menus.all().select_related("menu").order_by(SORT_ORDER)
        menu_names = [item.menu.name for item in sorted_diet_menus]

        if plan.meal_type == "조식":
            meal_map[day]["breakfast"] = menu_names
        elif plan.meal_type == "중식":
            meal_map[day]["lunch"] = menu_names
        elif plan.meal_type == "석식":
            meal_map[day]["dinner"] = menu_names

    # 달력 구조 생성 (일~토)
    calendar_weeks = []
    month_calendar = calendar.Calendar(firstweekday=calendar.SUNDAY)

    for week in month_calendar.monthdayscalendar(selected_year, selected_month):
        week_cells = []
        for day_num in week:
            if day_num == 0:
                week_cells.append(None)
            else:
                week_cells.append(
                    {
                        "day": day_num,
                        "meals": meal_map.get(day_num, {"breakfast": [], "lunch": [], "dinner": []}),
                    }
                )
        calendar_weeks.append(week_cells)

    context = {
        "selected_year": selected_year,
        "selected_month": selected_month,
        "years": range(today.year - 1, today.year + 2),
        "months": range(1, 13),
        "calendar_weeks": calendar_weeks,
    }

    return render(request, "docs/monthly_report.html", context)


def export_weekly_mealplan_excel(request):
    """
    주간 식단표(8일치) 데이터를 엑셀 파일로 내보냅니다.
    """
    # 1. 날짜 파라미터 처리
    target_date_str = request.GET.get("date") or request.GET.get("start_date")
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    start_date = target_date - timedelta(days=target_date.weekday())
    dates = [start_date + timedelta(days=i) for i in range(8)]

    # 2. 데이터 조회
    all_plans = (
        DietPlan.objects.filter(
            target_date__gte=dates[0],
            target_date__lte=dates[7],
        )
        .prefetch_related("diet_menus__menu")
    )

    weekly_data = {d: {"조식": [], "중식": [], "석식": []} for d in dates}
    for plan_obj in all_plans:
        if plan_obj.target_date in weekly_data:
            sorted_diet_menus = plan_obj.diet_menus.all().select_related("menu").order_by(SORT_ORDER)
            menus = [dm.menu.name for dm in sorted_diet_menus]
            weekly_data[plan_obj.target_date][plan_obj.meal_type] = menus

    # 3. 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = f"주간식단표_{start_date.strftime('%m%d')}"

    # 스타일 설정
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)

    # 4. 헤더 작성 (가로축: 월~일)
    weekdays_list = ["월", "화", "수", "목", "금", "토", "일"]
    ws.cell(row=1, column=1, value="구분").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = border

    for i, (d, w) in enumerate(zip(dates[:7], weekdays_list)):
        cell = ws.cell(row=1, column=i + 2, value=f"{d.strftime('%m/%d')}({w})")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # 컬럼 너비 조정
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 20

    # 5. 데이터 작성 (세로축: 조식, 중식, 석식, 비고)
    meal_types = ["조식", "중식", "석식", "비고"]
    for r_idx, meal_type in enumerate(meal_types):
        row_num = r_idx + 2
        label_cell = ws.cell(row=row_num, column=1, value=meal_type)
        label_cell.font = header_font
        label_cell.fill = header_fill
        label_cell.alignment = center_align
        label_cell.border = border
        ws.row_dimensions[row_num].height = 60

        for c_idx in range(7):
            d = dates[c_idx]
            col_num = c_idx + 2
            
            if meal_type == "비고":
                if c_idx == 6: # 일요일
                    next_monday = dates[7]
                    val = "[다음주 월요일 조식]\n" + "\n".join(weekly_data[next_monday].get("조식", []))
                    cell = ws.cell(row=row_num, column=col_num, value=val)
                    cell.font = Font(color="0000FF") # 파란색
                else:
                    cell = ws.cell(row=row_num, column=col_num, value="-")
            else:
                menus = weekly_data[d].get(meal_type, [])
                cell = ws.cell(row=row_num, column=col_num, value="\n".join(menus))
            
            cell.alignment = center_align
            cell.border = border

    # 6. HTTP 응답 반환
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"Weekly_MealPlan_{start_date.strftime('%Y-%m-%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
