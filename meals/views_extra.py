
import csv
import io
from datetime import datetime, date

import openpyxl
from django.contrib import messages
from django.db import transaction
from django.shortcuts import redirect, render

from .models import DietMenu, DietPlan, Menu

MEAL_LABELS = {
    "조식": "조식",
    "breakfast": "조식",
    "아침": "조식",
    "중식": "중식",
    "lunch": "중식",
    "점심": "중식",
    "석식": "석식",
    "dinner": "석식",
    "저녁": "석식",
}


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _compact(value):
    return _clean(value).replace(" ", "").lower()


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    for fmt in ["%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"]:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _parse_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(str(value).replace(",", "").strip()))
    except Exception:
        return default


def _split_menus(value):
    text = _clean(value)
    if not text:
        return []
    for sep in ["/", "|", "\n", ";"]:
        text = text.replace(sep, ",")
    return [item.strip() for item in text.split(",") if item.strip()]


def _meal_type(value):
    return MEAL_LABELS.get(_compact(value), MEAL_LABELS.get(_clean(value), ""))


def _read_upload_rows(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        raw = uploaded_file.read()
        for enc in ["utf-8-sig", "cp949", "euc-kr"]:
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active

    # 상단 안내문이 있어도 헤더를 찾는다.
    header_row = None
    headers = []
    for r in range(1, min(sheet.max_row, 20) + 1):
        values = [_clean(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
        compact = [_compact(v) for v in values]
        if any(v in compact for v in ["날짜", "일자", "date"]) and (
            any(v in compact for v in ["끼니", "mealtype"]) or any(v in compact for v in ["조식", "중식", "석식"])
        ):
            header_row = r
            headers = values
            break
    if header_row is None:
        header_row = 1
        headers = [_clean(sheet.cell(1, c).value) for c in range(1, sheet.max_column + 1)]

    rows = []
    for r in range(header_row + 1, sheet.max_row + 1):
        row = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            value = sheet.cell(r, c).value
            if value not in [None, ""]:
                empty = False
            row[h] = value
        if not empty:
            rows.append(row)
    return rows


def _find(row, aliases):
    compact_map = {_compact(k): v for k, v in row.items()}
    for alias in aliases:
        key = _compact(alias)
        if key in compact_map:
            return compact_map[key]
    return None


def _collect_mealplan_rows(rows):
    """두 가지 형식 모두 지원.

    1) 행 단위: 날짜 / 끼니 / 메뉴명 / 기준식수
    2) 주간형: 날짜 / 조식 / 중식 / 석식 / 기준식수_조식 ...
    """
    grouped = {}
    skipped = []
    for idx, row in enumerate(rows, start=2):
        target_date = _parse_date(_find(row, ["날짜", "일자", "date", "식단일"]))
        if not target_date:
            skipped.append({"row": idx, "reason": "날짜를 읽을 수 없음"})
            continue

        meal = _meal_type(_find(row, ["끼니", "meal_type", "mealtype", "구분"]))
        menu_name = _find(row, ["메뉴명", "메뉴", "menu", "음식명"])
        headcount = _parse_int(_find(row, ["기준식수", "식수", "headcount", "인원"]), 0)

        if meal and menu_name:
            key = (target_date, meal)
            grouped.setdefault(key, {"headcount": headcount, "menus": []})
            grouped[key]["headcount"] = headcount or grouped[key]["headcount"]
            grouped[key]["menus"].extend(_split_menus(menu_name))
            continue

        # wide format
        found_any = False
        for label in ["조식", "중식", "석식"]:
            menus = _split_menus(_find(row, [label]))
            if not menus:
                continue
            found_any = True
            hc = _parse_int(_find(row, [f"기준식수_{label}", f"{label}식수", f"식수_{label}"]), 0)
            key = (target_date, label)
            grouped.setdefault(key, {"headcount": hc, "menus": []})
            grouped[key]["headcount"] = hc or grouped[key]["headcount"]
            grouped[key]["menus"].extend(menus)
        if not found_any:
            skipped.append({"row": idx, "reason": "끼니/메뉴 정보를 찾을 수 없음"})
    return grouped, skipped


@transaction.atomic
def _save_grouped_mealplans(grouped):
    plan_count = 0
    menu_count = 0
    for (target_date, meal_type), payload in grouped.items():
        plan, _ = DietPlan.objects.update_or_create(
            target_date=target_date,
            meal_type=meal_type,
            defaults={"headcount": payload.get("headcount") or 0},
        )
        DietMenu.objects.filter(diet_plan=plan).delete()
        seen = set()
        for menu_name in payload.get("menus", []):
            clean_name = _clean(menu_name)
            if not clean_name or clean_name in seen:
                continue
            seen.add(clean_name)
            menu, _ = Menu.objects.get_or_create(name=clean_name, defaults={"category": "기타"})
            DietMenu.objects.get_or_create(diet_plan=plan, menu=menu)
            menu_count += 1
        plan_count += 1
    return plan_count, menu_count


def monthly_mealplan_create(request):
    if request.method == "POST" and request.FILES.get("mealplan_file"):
        return mealplan_bulk_upload(request)
    if request.method == "POST":
        messages.info(request, "월간 식단은 엑셀 업로드 또는 주/일간 식단 등록 화면에서 입력해 주세요.")
        return redirect("meals:mealplan_list")
    return render(request, "meals/monthly_mealplan_create.html")


def mealplan_bulk_upload(request):
    if request.method == "GET":
        return render(request, "meals/mealplan_bulk_upload.html")

    uploaded_file = request.FILES.get("mealplan_file") or request.FILES.get("file")
    if not uploaded_file:
        messages.error(request, "CSV 또는 XLSX 파일을 선택해 주세요.")
        return redirect("meals:mealplan_bulk_upload")
    if not uploaded_file.name.lower().endswith((".xlsx", ".xlsm", ".csv")):
        messages.error(request, "현재는 .xlsx, .xlsm, .csv 파일만 지원합니다.")
        return redirect("meals:mealplan_bulk_upload")

    try:
        rows = _read_upload_rows(uploaded_file)
        grouped, skipped = _collect_mealplan_rows(rows)
        if not grouped:
            messages.error(request, "저장할 식단 데이터를 찾지 못했습니다. 날짜/끼니/메뉴명 컬럼을 확인해 주세요.")
            return redirect("meals:mealplan_bulk_upload")
        plan_count, menu_count = _save_grouped_mealplans(grouped)
        if skipped:
            messages.warning(request, f"일부 행 {len(skipped)}건은 제외되었습니다.")
        messages.success(request, f"식단 엑셀 업로드 완료: 식단 {plan_count}건, 메뉴 연결 {menu_count}건 반영")
        return redirect("meals:mealplan_list")
    except Exception as exc:
        messages.error(request, f"식단 엑셀 업로드 중 오류가 발생했습니다: {exc}")
        return redirect("meals:mealplan_bulk_upload")
