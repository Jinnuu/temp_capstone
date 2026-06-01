from datetime import date, datetime, timedelta
import calendar
import traceback
from datetime import date
from urllib.parse import urlencode

from django.contrib import messages
from django.db.models import Case, IntegerField, Value, When
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill

from .forms import AttendancePredictionForm, PredictionFilterForm
from .models import AttendancePrediction
from .services.ingredient_calc import get_all_day_requirements
from .services.prediction_service import run_attendance_prediction
from meals.models import DietPlan
from meals.models import DietPlan, DietMenu, Menu, Recipe

MEAL_LABELS = {
    AttendancePrediction.MealType.BREAKFAST: "조식",
    AttendancePrediction.MealType.LUNCH: "중식",
    AttendancePrediction.MealType.DINNER: "석식",
}


def forecasting_home(request):
    return redirect("forecasting:prediction_create")


def prediction_create_view(request):
    if request.method == "POST":
        form = AttendancePredictionForm(request.POST)

        if form.is_valid():
            prediction_date = form.cleaned_data["prediction_date"]
            meal_types = [
                AttendancePrediction.MealType.BREAKFAST,
                AttendancePrediction.MealType.LUNCH,
                AttendancePrediction.MealType.DINNER,
            ]

            success_count = 0
            failed_meals = []

            for meal_type in meal_types:
                try:
                    run_attendance_prediction(
                        prediction_date=prediction_date,
                        meal_type=meal_type,
                        menu_text=None,
                    )
                    success_count += 1
                except Exception as e:
                    traceback.print_exc()
                    failed_meals.append(f"{meal_type}({e})")

            if success_count:
                messages.success(
                    request,
                    f"{prediction_date} 기준 예측 {success_count}건이 생성되었습니다.",
                )

                if failed_meals:
                    messages.warning(
                        request,
                        "일부 끼니 예측에 실패했습니다: " + ", ".join(failed_meals),
                    )

                url = reverse("forecasting:prediction_list")
                query_string = urlencode({"target_date": prediction_date})
                return redirect(f"{url}?{query_string}")

            messages.error(request, "예측 실행에 실패했습니다. 로그를 확인해주세요.")

            if failed_meals:
                messages.warning(request, "실패한 끼니: " + ", ".join(failed_meals))
    else:
        initial = {}
        selected_date = request.GET.get("date")

        if selected_date:
            initial["prediction_date"] = selected_date

        form = AttendancePredictionForm(initial=initial)

    return render(request, "forecasting/predict_form.html", {"form": form})


def prediction_calendar_view(request):
    today = date.today()
    selected_year = int(request.GET.get("year", today.year))
    selected_month = int(request.GET.get("month", today.month))

    cal = calendar.Calendar(firstweekday=6)
    month_days = cal.monthdatescalendar(selected_year, selected_month)

    predictions = (
        AttendancePrediction.objects.filter(
            prediction_date__year=selected_year,
            prediction_date__month=selected_month,
        )
        .order_by("prediction_date", "meal_type", "-created_at")
    )

    prediction_map = {}

    for prediction in predictions:
        day_map = prediction_map.setdefault(prediction.prediction_date, {})

        if prediction.meal_type not in day_map:
            day_map[prediction.meal_type] = prediction.predicted_count

    calendar_weeks = []

    for week in month_days:
        row = []

        for day in week:
            if day.month != selected_month:
                row.append(None)
                continue

            day_predictions = prediction_map.get(day, {})
            ordered_meal_types = [
                AttendancePrediction.MealType.BREAKFAST,
                AttendancePrediction.MealType.LUNCH,
                AttendancePrediction.MealType.DINNER,
            ]

            summary_parts = []

            for meal_type in ordered_meal_types:
                count = day_predictions.get(meal_type)

                if count is not None:
                    summary_parts.append(f"{MEAL_LABELS.get(meal_type, meal_type)} {count}명")

            total_count = sum(day_predictions.values()) if day_predictions else None

            row.append(
                {
                    "date": day,
                    "day": day.day,
                    "is_today": day == today,
                    "has_prediction": bool(day_predictions),
                    "prediction_summary": " · ".join(summary_parts),
                    "total_count": total_count,
                }
            )

        calendar_weeks.append(row)

    context = {
        "selected_year": selected_year,
        "selected_month": selected_month,
        "months": range(1, 13),
        "calendar_weeks": calendar_weeks,
    }

    return render(request, "forecasting/prediction_calendar.html", context)


def prediction_result_view(request, prediction_id):
    prediction = get_object_or_404(AttendancePrediction, id=prediction_id)
    return render(request, "forecasting/predict_result.html", {"prediction": prediction})


def prediction_list_view(request):
    form = PredictionFilterForm(request.GET or None)

    predictions = AttendancePrediction.objects.annotate(
        meal_order=Case(
            When(meal_type=AttendancePrediction.MealType.BREAKFAST, then=Value(1)),
            When(meal_type=AttendancePrediction.MealType.LUNCH, then=Value(2)),
            When(meal_type=AttendancePrediction.MealType.DINNER, then=Value(3)),
            default=Value(99),
            output_field=IntegerField(),
        )
    )

    if form.is_valid():
        target_date = form.cleaned_data.get("target_date")
        start_date = form.cleaned_data.get("start_date")
        end_date = form.cleaned_data.get("end_date")
        order = form.cleaned_data.get("order") or "latest"

        if target_date:
            predictions = predictions.filter(prediction_date=target_date)
        else:
            if start_date:
                predictions = predictions.filter(prediction_date__gte=start_date)
            if end_date:
                predictions = predictions.filter(prediction_date__lte=end_date)

        if order == "oldest":
            predictions = predictions.order_by("prediction_date", "meal_order", "created_at")
        else:
            predictions = predictions.order_by("-prediction_date", "meal_order", "-created_at")
    else:
        predictions = predictions.order_by("-prediction_date", "meal_order", "-created_at")

    return render(
        request,
        "forecasting/prediction_list.html",
        {
            "predictions": predictions,
            "form": form,
        },
    )


def ingredient_requirement_view(request):
    selected_date = request.GET.get("date", date.today().isoformat())
    requirements = get_all_day_requirements(selected_date)

    context = {
        "requirements": requirements,
        "selected_date": selected_date,
        "is_all_day": True,
    }

    return render(request, "forecasting/required_ingredient_list.html", context)


def get_monday_sunday_range(date_str):
    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    monday = target_date - timedelta(days=target_date.weekday())
    sunday = monday + timedelta(days=6)
    return target_date, monday, sunday


def weekly_forecast_report_view(request):
    target_date, monday, sunday = get_monday_sunday_range(request.GET.get("date"))

    meal_mapping = {
        "breakfast": "조식",
        "lunch": "중식",
        "dinner": "석식",
    }
    
    # 1. 대상 기간의 식단(DietPlan) 조회
    diet_plans = DietPlan.objects.filter(
        target_date__range=[monday, sunday]
    ).prefetch_related("diet_menus__menu").order_by("target_date")

    # 2. 대상 기간의 예측(AttendancePrediction) 조회
    predictions = AttendancePrediction.objects.filter(
        prediction_date__range=[monday, sunday]
    ).order_by("prediction_date")

    # 3. 데이터 매핑 (날짜/끼니 기준)
    prediction_map = {}
    for p in predictions:
        meal_ko = meal_mapping.get(p.meal_type)
        if meal_ko:
            prediction_map[(p.prediction_date, meal_ko)] = p.predicted_count

    report_items = []
    
    current_date = monday
    while current_date <= sunday:
        day_plans = diet_plans.filter(target_date=current_date)
        meal_order = {"조식": 1, "중식": 2, "석식": 3}
        day_plans_list = sorted(list(day_plans), key=lambda x: meal_order.get(x.meal_type, 99))
        
        day_span = len(day_plans_list) if day_plans_list else 1
        is_first_of_day = True
        
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        weekday_str = weekdays[current_date.weekday()]

        if not day_plans_list:
            # 식단이 없는 날짜도 표시할 수 있도록 빈 항목 추가 (선택사항)
            pass
        
        for plan in day_plans_list:
            predicted = prediction_map.get((current_date, plan.meal_type), 0)
            actual = plan.headcount
            
            error_rate = 0
            if actual and actual > 0:
                error_rate = ((predicted - actual) / actual) * 100
            
            menus = [dm.menu.name for dm in plan.diet_menus.all()]
            
            report_items.append({
                "date": current_date,
                "weekday": weekday_str,
                "meal_type": plan.meal_type,
                "menus": menus,
                "predicted_count": predicted,
                "actual_count": actual,
                "error_rate": error_rate,
                "is_first_of_day": is_first_of_day,
                "day_span": day_span,
            })
            is_first_of_day = False
            
        current_date += timedelta(days=1)

    context = {
        "target_date": target_date,
        "monday": monday,
        "sunday": sunday,
        "prev_week_date": (monday - timedelta(days=7)).isoformat(),
        "next_week_date": (monday + timedelta(days=7)).isoformat(),
        "report_items": report_items,
    }
    return render(request, "docs/weekly_forecast_report.html", context)


def export_weekly_forecast_excel(request):
    target_date, monday, sunday = get_monday_sunday_range(request.GET.get("date"))

    meal_mapping = {
        "breakfast": "조식",
        "lunch": "중식",
        "dinner": "석식",
    }

    diet_plans = DietPlan.objects.filter(
        target_date__range=[monday, sunday]
    ).prefetch_related("diet_menus__menu").order_by("target_date")

    predictions = AttendancePrediction.objects.filter(
        prediction_date__range=[monday, sunday]
    )

    prediction_map = {}
    for p in predictions:
        meal_ko = meal_mapping.get(p.meal_type)
        if meal_ko:
            prediction_map[(p.prediction_date, meal_ko)] = p.predicted_count

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주간예측보고서"

    headers = ["날짜", "요일", "끼니", "식단(메뉴명)", "예측 식수", "실제 식수", "오차율(%)"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True)
    header_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_date = monday
    while current_date <= sunday:
        day_plans = diet_plans.filter(target_date=current_date)
        meal_order = {"조식": 1, "중식": 2, "석식": 3}
        day_plans_list = sorted(list(day_plans), key=lambda x: meal_order.get(x.meal_type, 99))
        
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        weekday_str = weekdays[current_date.weekday()]

        for plan in day_plans_list:
            predicted = prediction_map.get((current_date, plan.meal_type), 0)
            actual = plan.headcount
            error_rate = ""
            if actual and actual > 0:
                error_rate = f"{((predicted - actual) / actual) * 100:.1f}%"
            elif actual == 0 or actual is None:
                error_rate = "-"
                
            menus = ", ".join([dm.menu.name for dm in plan.diet_menus.all()])
            
            ws.append([
                current_date.strftime("%Y-%m-%d"),
                weekday_str,
                plan.meal_type,
                menus,
                predicted,
                actual if actual is not None else 0,
                error_rate
            ])
        current_date += timedelta(days=1)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row[0].row, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"weekly_forecast_{monday}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

