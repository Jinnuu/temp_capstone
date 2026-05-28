from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Case, DateField, F, IntegerField, Sum, Value, When
from django.db.models.functions import Cast, Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

from forecasting.models import AttendancePrediction
from forecasting.services.ingredient_calc import get_all_day_requirements
from inventory.models import Ingredient, InventoryLog
from meals.models import DietPlan
from .models import OrderItem, PurchaseOrder


CATEGORY_ORDER = ["밥", "국", "주반찬", "부반찬", "김치", "간식"]

SORT_ORDER = Case(
    *[When(menu__category=cat, then=Value(pos)) for pos, cat in enumerate(CATEGORY_ORDER)],
    default=Value(len(CATEGORY_ORDER)),
    output_field=IntegerField(),
)

MEAL_TO_PREDICTION_KEY = {
    "조식": "breakfast",
    "중식": "lunch",
    "석식": "dinner",
}


def to_decimal(value, default="0"):
    try:
        return Decimal(str(value if value is not None else default))
    except Exception:
        return Decimal(default)


def get_or_create_dummy_user(request):
    if request.user.is_authenticated:
        return request.user

    User = get_user_model()
    user = User.objects.first()

    if not user:
        user = User.objects.create_user(username="admin_dummy", password="dummy_password")

    return user


def get_current_stock(ingredient):
    """InventoryLog 기준 현재고 계산. 기존 데이터가 영문/한글 로그값을 섞어 쓰는 경우까지 방어."""
    in_sum = (
        InventoryLog.objects.filter(
            ingredient=ingredient,
            log_type__in=[InventoryLog.LogType.IN, "IN", "입고", "1"],
        ).aggregate(total=Sum("quantity"))["total"]
        or Decimal("0")
    )
    out_sum = (
        InventoryLog.objects.filter(
            ingredient=ingredient,
            log_type__in=[InventoryLog.LogType.OUT, "OUT", "출고", "2"],
        ).aggregate(total=Sum("quantity"))["total"]
        or Decimal("0")
    )
    waste_sum = (
        InventoryLog.objects.filter(
            ingredient=ingredient,
            log_type__in=[InventoryLog.LogType.WASTE, "WASTE", "폐기", "3"],
        ).aggregate(total=Sum("quantity"))["total"]
        or Decimal("0")
    )
    adj_sum = (
        InventoryLog.objects.filter(
            ingredient=ingredient,
            log_type__in=[InventoryLog.LogType.ADJ, "ADJ", "조정", "4"],
        ).aggregate(total=Sum("quantity"))["total"]
        or Decimal("0")
    )

    return to_decimal(in_sum) - to_decimal(out_sum) - to_decimal(waste_sum) + to_decimal(adj_sum)


def get_serving_count(plan):
    """예측 식수 우선, 없으면 MealForecast, 그것도 없으면 DietPlan.headcount 사용."""
    prediction_key = MEAL_TO_PREDICTION_KEY.get(plan.meal_type)

    if prediction_key:
        prediction = (
            AttendancePrediction.objects.filter(
                prediction_date=plan.target_date,
                meal_type=prediction_key,
            )
            .order_by("-created_at")
            .first()
        )

        if prediction:
            return int(prediction.predicted_count or 0), "예측"

    forecast = getattr(plan, "forecast", None)

    if forecast:
        return int(forecast.predicted_count or 0), "예측"

    return int(plan.headcount or 0), "식단 기준"


def order_create(request):
    """간편주문 생성. GET 파라미터 또는 날짜 기준 부족분을 수량에 자동 채운다."""
    if request.method == "POST":
        ingredient_ids = request.POST.getlist("ingredient_id[]")
        quantities = request.POST.getlist("quantity[]")
        target_date_val = request.POST.get("target_date")

        if ingredient_ids and quantities:
            user = get_or_create_dummy_user(request)
            po = PurchaseOrder.objects.create(
                supplier=user,
                status=PurchaseOrder.Status.PENDING,
            )

            total_amount = 0

            for ing_id, qty in zip(ingredient_ids, quantities):
                q = to_decimal(qty)

                if q <= 0:
                    continue

                ing = get_object_or_404(Ingredient, id=ing_id)
                estimated = int(q * to_decimal(ing.unit_price))
                total_amount += estimated

                OrderItem.objects.create(
                    purchase_order=po,
                    ingredient=ing,
                    target_date=target_date_val if target_date_val else None,
                    required_qty=q,
                    missing_qty=q,
                    order_unit_price=ing.unit_price,
                    estimated_price=estimated,
                )

            if total_amount <= 0:
                po.delete()
                messages.error(request, "발주할 품목이 없습니다.")
                return redirect("procurement:order_create")

            po.total_amount = total_amount
            po.save()

            messages.success(request, f"발주서 #{po.id} 건이 성공적으로 전송되었습니다.")
            return redirect("procurement:order_detail", pk=po.id)

    items_names = request.GET.getlist("items")
    amounts = request.GET.getlist("amounts")
    shortage_map = dict(zip(items_names, amounts))

    target_date = request.GET.get("target_date")

    if target_date:
        requirements = get_all_day_requirements(target_date)

        for name, data in requirements.get("items", {}).items():
            order_amount = data.get("order_amount", 0)

            if order_amount and float(order_amount) > 0:
                shortage_map[name] = str(order_amount)

    ingredients = Ingredient.objects.all().order_by("name")
    categories = (
        Ingredient.objects.exclude(category__isnull=True)
        .exclude(category__exact="")
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )

    return render(
        request,
        "procurement/order_create.html",
        {
            "ingredients": ingredients,
            "categories": categories,
            "shortage_map": shortage_map,
            "target_date": target_date or date.today().isoformat(),
        },
    )


def order_list(request):
    orders = PurchaseOrder.objects.all().order_by("-created_at")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    status_filter = request.GET.get("status")

    if start_date:
        orders = orders.filter(created_at__date__gte=start_date)

    if end_date:
        orders = orders.filter(created_at__date__lte=end_date)

    if status_filter and status_filter != "전체":
        orders = orders.filter(status=status_filter)

    orders_by_supplier = defaultdict(list)

    for order in orders:
        orders_by_supplier[order.supplier].append(order)

    supplier_groups = []

    for supplier, s_orders in orders_by_supplier.items():
        supplier_groups.append(
            {
                "supplier": supplier,
                "orders": s_orders,
            }
        )

    supplier_groups.sort(key=lambda x: x["supplier"].name or x["supplier"].username)

    return render(
        request,
        "procurement/order_list.html",
        {
            "orders": orders,
            "supplier_groups": supplier_groups,
            "start_date": start_date,
            "end_date": end_date,
            "status_filter": status_filter,
        },
    )


def order_detail(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)

    meal_order = Case(
        When(meal_type="조식", then=Value(1)),
        When(meal_type="중식", then=Value(2)),
        When(meal_type="석식", then=Value(3)),
        default=Value(4),
        output_field=IntegerField(),
    )

    items = order.items.select_related("ingredient").order_by(
        "target_date",
        meal_order,
        "ingredient__name",
    )

    subtotal = order.total_amount
    vat = int(subtotal * 0.1)
    total = subtotal + vat

    return render(
        request,
        "procurement/order_detail.html",
        {
            "order": order,
            "items": items,
            "subtotal": subtotal,
            "vat": vat,
            "total": total,
        },
    )


def order_status_update(request, pk):
    if request.method == "POST":
        order = get_object_or_404(PurchaseOrder, pk=pk)
        new_status = request.POST.get("status")

        if new_status in ["배송중", "배송완료", "완료"]:
            final_status = "완료"
        else:
            final_status = new_status

        valid_choices = [choice[0] for choice in PurchaseOrder.Status.choices]

        if final_status in valid_choices:
            if final_status == "완료" and order.status != "완료":
                for item in order.items.all():
                    InventoryLog.objects.create(
                        ingredient=item.ingredient,
                        log_type=InventoryLog.LogType.IN,
                        quantity=item.required_qty,
                        description=f"발주 #{order.id} 배송 완료로 인한 자동 입고",
                    )

                messages.success(request, "배송 완료 및 실시간 입고 재고 반영이 완료되었습니다.")

            order.status = final_status
            order.save()
            messages.success(request, "발주 상태가 변경되었습니다.")
        else:
            messages.error(request, f"유효하지 않은 상태 값입니다: {new_status}")

    return redirect("procurement:order_detail", pk=pk)


def order_from_mealplan(request):
    """식단, 예측 식수, 레시피, 현재고, 안전재고를 기준으로 부족 수량을 자동 계산해 발주한다."""
    if request.method == "POST":
        ingredient_ids = request.POST.getlist("ingredient_id[]")
        quantities = request.POST.getlist("quantity[]")
        target_dates = request.POST.getlist("target_date[]")
        meal_types = request.POST.getlist("meal_type[]")

        order_data = {}

        for ing_id_str, qty_str, target_date_str, meal_type_str in zip(
            ingredient_ids,
            quantities,
            target_dates,
            meal_types,
        ):
            try:
                ing_id = int(ing_id_str)
            except Exception:
                continue

            q = to_decimal(qty_str)

            if q <= 0:
                continue

            ing = get_object_or_404(Ingredient, id=ing_id)

            if not ing.supplier:
                continue

            po_key = (ing.supplier_id, target_date_str or None)

            if po_key not in order_data:
                order_data[po_key] = {}

            item_key = (ing_id, meal_type_str or None)

            if item_key not in order_data[po_key]:
                order_data[po_key][item_key] = Decimal("0")

            order_data[po_key][item_key] += q

        if not order_data:
            messages.error(request, "발주할 품목이 없습니다.")
            return redirect("procurement:order_from_mealplan")

        with transaction.atomic():
            created_po_ids = []

            for (supplier_id, target_date_str), items_dict in order_data.items():
                supplier_user = get_user_model().objects.get(id=supplier_id)
                po = PurchaseOrder.objects.create(
                    supplier=supplier_user,
                    status=PurchaseOrder.Status.PENDING,
                    total_amount=0,
                )

                po_total = 0

                for (ing_id, meal_type_str), q in items_dict.items():
                    ing = Ingredient.objects.get(id=ing_id)
                    subtotal = int(q * to_decimal(ing.unit_price))
                    po_total += subtotal

                    OrderItem.objects.create(
                        purchase_order=po,
                        ingredient=ing,
                        target_date=target_date_str if target_date_str else None,
                        meal_type=meal_type_str,
                        required_qty=q,
                        missing_qty=q,
                        order_unit_price=ing.unit_price,
                        estimated_price=subtotal,
                    )

                po.total_amount = po_total
                po.save()
                created_po_ids.append(po.id)

        messages.success(request, f"총 {len(created_po_ids)}건의 업체별 발주서가 성공적으로 생성되었습니다.")
        return redirect("procurement:order_list")

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = date.today()
    else:
        start_date = date.today()

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            end_date = start_date + timedelta(days=6)
    else:
        end_date = start_date + timedelta(days=6)

    meal_order = Case(
        When(meal_type="조식", then=Value(1)),
        When(meal_type="중식", then=Value(2)),
        When(meal_type="석식", then=Value(3)),
        default=Value(4),
        output_field=IntegerField(),
    )

    plans = (
        DietPlan.objects.filter(
            target_date__gte=start_date,
            target_date__lte=end_date,
        )
        .prefetch_related("diet_menus__menu__recipes__ingredient__supplier")
        .order_by("target_date", meal_order)
    )

    # 기간 내에서 같은 식재료가 여러 번 쓰일 때 현재고를 순차 차감해 중복 과소/과대 계산을 줄인다.
    stock_balance = {
        ingredient.id: get_current_stock(ingredient)
        for ingredient in Ingredient.objects.all()
    }

    analysis_data = {}

    for plan in plans:
        dt = plan.target_date

        if dt not in analysis_data:
            analysis_data[dt] = []

        serving_count, serving_source = get_serving_count(plan)

        plan_item = {
            "meal_type": plan.meal_type,
            "headcount": plan.headcount,
            "serving_count": serving_count,
            "serving_source": serving_source,
            "menus": [],
        }

        sorted_diet_menus = plan.diet_menus.all().select_related("menu").order_by(SORT_ORDER)

        for diet_menu in sorted_diet_menus:
            menu = diet_menu.menu
            menu_item = {
                "id": menu.id,
                "name": menu.name,
                "ingredients": [],
            }

            for recipe in menu.recipes.all():
                ing = recipe.ingredient
                current_before = stock_balance.get(ing.id, Decimal("0"))
                required_per_person = to_decimal(recipe.required_amount)
                needed_qty = to_decimal(serving_count) * required_per_person
                safe_stock = to_decimal(ing.safe_stock_level)
                shortage_qty = needed_qty + safe_stock - current_before
                order_qty = shortage_qty if shortage_qty > 0 else Decimal("0")

                # 현재 끼니 소비량만큼 재고를 차감한다. 음수는 0으로 고정한다.
                stock_balance[ing.id] = max(Decimal("0"), current_before - needed_qty)

                menu_item["ingredients"].append(
                    {
                        "id": ing.id,
                        "name": ing.name,
                        "unit": ing.unit,
                        "price": ing.unit_price,
                        "category": ing.category,
                        "spec": ing.spec,
                        "description": ing.description,
                        "supplier_name": ing.supplier.name if ing.supplier else "미정",
                        "current_stock": round(current_before, 2),
                        "safe_stock": round(safe_stock, 2),
                        "needed_qty": round(needed_qty, 2),
                        "order_qty": round(order_qty, 2),
                    }
                )

            plan_item["menus"].append(menu_item)

        analysis_data[dt].append(plan_item)

    sorted_analysis = [
        {
            "date": dt,
            "plans": analysis_data[dt],
        }
        for dt in sorted(analysis_data.keys())
    ]

    return render(
        request,
        "procurement/order_from_mealplan.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "analysis_data": sorted_analysis,
            "all_ingredients": Ingredient.objects.all().order_by("name"),
        },
    )


def get_monday_sunday_range(date_str):
    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    monday = target_date - timedelta(days=target_date.weekday())
    sunday = monday + timedelta(days=6)
    return target_date, monday, sunday


def weekly_procurement_report(request):
    target_date, monday, sunday = get_monday_sunday_range(request.GET.get("date"))

    # Use Coalesce to fallback to purchase_order__created_at__date if target_date is NULL
    order_items = OrderItem.objects.annotate(
        display_date=Coalesce(F("target_date"), Cast(F("purchase_order__created_at"), DateField()))
    ).filter(
        display_date__range=[monday, sunday]
    ).select_related("ingredient", "purchase_order__supplier").order_by("display_date", "ingredient__name")

    # Group by date
    analysis_data = defaultdict(list)
    for item in order_items:
        analysis_data[item.display_date].append(item)

    report_data = [
        {"date": dt, "items": analysis_data[dt]}
        for dt in sorted(analysis_data.keys())
    ]

    context = {
        "report_data": report_data,
        "monday": monday,
        "sunday": sunday,
        "target_date": target_date,
        "prev_week_date": (monday - timedelta(days=7)).isoformat(),
        "next_week_date": (monday + timedelta(days=7)).isoformat(),
    }
    return render(request, "docs/weekly_procurement_report.html", context)


def vendor_procurement_report(request):
    target_date, monday, sunday = get_monday_sunday_range(request.GET.get("date"))
    supplier_id = request.GET.get("supplier_id")

    suppliers = get_user_model().objects.filter(purchase_orders__isnull=False).distinct()
    selected_supplier = None
    report_data = []

    if supplier_id:
        selected_supplier = get_object_or_404(get_user_model(), id=supplier_id)
        order_items = OrderItem.objects.annotate(
            display_date=Coalesce(F("target_date"), Cast(F("purchase_order__created_at"), DateField()))
        ).filter(
            display_date__range=[monday, sunday],
            purchase_order__supplier=selected_supplier
        ).select_related("ingredient", "purchase_order").order_by("display_date", "ingredient__name")

        analysis_data = defaultdict(list)
        for item in order_items:
            analysis_data[item.display_date].append(item)

        report_data = [
            {"date": dt, "items": analysis_data[dt]}
            for dt in sorted(analysis_data.keys())
        ]

    context = {
        "suppliers": suppliers,
        "selected_supplier": selected_supplier,
        "current_supplier_id": supplier_id,
        "report_data": report_data,
        "monday": monday,
        "sunday": sunday,
        "target_date": target_date,
        "prev_week_date": (monday - timedelta(days=7)).isoformat(),
        "next_week_date": (monday + timedelta(days=7)).isoformat(),
    }
    return render(request, "docs/vendor_procurement_report.html", context)


def export_weekly_procurement_excel(request):
    _, monday, sunday = get_monday_sunday_range(request.GET.get("date"))
    items = OrderItem.objects.annotate(
        display_date=Coalesce(F("target_date"), Cast(F("purchase_order__created_at"), DateField()))
    ).filter(
        display_date__range=[monday, sunday]
    ).select_related("ingredient", "purchase_order__supplier").order_by("display_date", "ingredient__name")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주간발주서"

    headers = ["No", "식단일", "분류", "식재료명", "규격", "단위", "단가", "소요량", "발주량", "금액", "공급처", "주문일자"]
    ws.append(headers)

    for i, item in enumerate(items, 1):
        ws.append([
            i,
            item.display_date.strftime("%Y-%m-%d") if item.display_date else "-",
            item.ingredient.category or "-",
            item.ingredient.name,
            item.ingredient.spec or "-",
            item.ingredient.unit,
            item.order_unit_price,
            float(item.required_qty),
            float(item.missing_qty),
            item.estimated_price,
            item.purchase_order.supplier.name or item.purchase_order.supplier.username,
            item.purchase_order.created_at.strftime("%Y-%m-%d")
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="weekly_procurement_{monday}.xlsx"'
    wb.save(response)
    return response


def export_vendor_procurement_excel(request):
    _, monday, sunday = get_monday_sunday_range(request.GET.get("date"))
    supplier_id = request.GET.get("supplier_id")
    
    if not supplier_id:
        return HttpResponse("업체를 선택해주세요.", status=400)

    supplier = get_object_or_404(get_user_model(), id=supplier_id)
    items = OrderItem.objects.annotate(
        display_date=Coalesce(F("target_date"), Cast(F("purchase_order__created_at"), DateField()))
    ).filter(
        display_date__range=[monday, sunday],
        purchase_order__supplier=supplier
    ).select_related("ingredient", "purchase_order").order_by("display_date", "ingredient__name")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"발주명세서_{supplier.name or supplier.username}"

    headers = ["No", "식단일", "분류", "식재료명", "규격", "단위", "단가", "소요량", "발주량", "금액", "주문일자"]
    ws.append(headers)

    for i, item in enumerate(items, 1):
        ws.append([
            i,
            item.display_date.strftime("%Y-%m-%d") if item.display_date else "-",
            item.ingredient.category or "-",
            item.ingredient.name,
            item.ingredient.spec or "-",
            item.ingredient.unit,
            item.order_unit_price,
            float(item.required_qty),
            float(item.missing_qty),
            item.estimated_price,
            item.purchase_order.created_at.strftime("%Y-%m-%d")
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"vendor_procurement_{supplier.name or supplier.username}_{monday}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def daily_inspection_report(request):
    date_str = request.GET.get("date")
    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    supplier_id = request.GET.get("supplier_id")
    suppliers = get_user_model().objects.filter(purchase_orders__isnull=False).distinct()
    selected_supplier = None
    report_items = []

    order_items = OrderItem.objects.annotate(
        display_date=Coalesce(F("target_date"), Cast(F("purchase_order__created_at"), DateField()))
    ).filter(
        display_date=target_date
    ).select_related("ingredient", "purchase_order__supplier").order_by("ingredient__category", "ingredient__name")

    if supplier_id:
        selected_supplier = get_object_or_404(get_user_model(), id=supplier_id)
        order_items = order_items.filter(purchase_order__supplier=selected_supplier)

    report_items = order_items
    total_amount = sum(item.estimated_price for item in report_items)

    context = {
        "target_date": target_date,
        "prev_day_date": (target_date - timedelta(days=1)).isoformat(),
        "next_day_date": (target_date + timedelta(days=1)).isoformat(),
        "suppliers": suppliers,
        "selected_supplier": selected_supplier,
        "current_supplier_id": supplier_id,
        "report_items": report_items,
        "total_amount": total_amount,
    }
    return render(request, "docs/daily_inspection_report.html", context)


def export_daily_inspection_excel(request):
    date_str = request.GET.get("date")
    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    supplier_id = request.GET.get("supplier_id")
    
    order_items = OrderItem.objects.annotate(
        display_date=Coalesce(F("target_date"), Cast(F("purchase_order__created_at"), DateField()))
    ).filter(
        display_date=target_date
    ).select_related("ingredient", "purchase_order__supplier").order_by("ingredient__category", "ingredient__name")

    filename_suffix = "전체"
    selected_supplier_name = "전체"
    if supplier_id:
        supplier = get_object_or_404(get_user_model(), id=supplier_id)
        order_items = order_items.filter(purchase_order__supplier=supplier)
        selected_supplier_name = supplier.name or supplier.username
        filename_suffix = selected_supplier_name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검수서_주문"

    # Set Column Widths
    col_widths = [10, 25, 15, 8, 10, 15, 10, 10, 8, 8, 12, 8, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Title
    ws.merge_cells("A1:M2")
    ws["A1"] = "검 수 서 (주 문)"
    ws["A1"].font = Font(size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Approval Table (Top Right)
    # Row 1: Label
    ws.merge_cells("J1:J2")
    ws["J1"] = "결재"
    ws["K1"] = "담당"
    ws["L1"] = "팀장"
    ws["M1"] = "점장"
    # Row 2: Empty for signatures
    # (J1:J2 merged carries the label)
    
    # Style Approval Label
    for cell_coord in ["J1", "K1", "L1", "M1"]:
        ws[cell_coord].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell_coord].font = Font(bold=True)
        ws[cell_coord].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws[cell_coord].fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Style Approval Content Boxes (Below)
    for row_idx in [2]:
        for col_letter in ["K", "L", "M"]:
            cell = ws[f"{col_letter}{row_idx}"]
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.row_dimensions[2].height = 40

    # Meta Info
    ws["A3"] = f"납품업체 : {selected_supplier_name}"
    ws["A4"] = "사업장 : S&P 급식소"
    ws["A5"] = f"입고일자 : {target_date}"
    for row in [3, 4, 5]:
        ws[f"A{row}"].font = Font(bold=True)

    # Table Headers (Double Row)
    # Row 6: Main Headers
    ws.merge_cells("A6:A7")
    ws["A6"] = "구분"
    ws.merge_cells("B6:B7")
    ws["B6"] = "품명"
    ws.merge_cells("C6:C7")
    ws["C6"] = "규격"
    ws.merge_cells("D6:D7")
    ws["D6"] = "단위"
    ws.merge_cells("E6:E7")
    ws["E6"] = "주문량"
    ws.merge_cells("F6:F7")
    ws["F6"] = "주문비고"
    ws.merge_cells("G6:G7")
    ws["G6"] = "원산지"
    
    ws.merge_cells("H6:M6")
    ws["H6"] = "검수 결과 (현장 수기 작성용)"
    
    # Row 7: Sub Headers
    ws["H7"] = "검수량"
    ws["I7"] = "포장"
    ws["J7"] = "품온"
    ws["K7"] = "소비기한"
    ws["L7"] = "품질"
    ws["M7"] = "조치"

    # Style Headers
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = openpyxl.styles.PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in [6, 7]:
        for col_idx in range(1, 14):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

    # Data Rows
    current_row = 8
    for item in order_items:
        ws.cell(row=current_row, column=1, value=item.ingredient.category or "-")
        ws.cell(row=current_row, column=2, value=item.ingredient.name)
        ws.cell(row=current_row, column=3, value=item.ingredient.spec or "-")
        ws.cell(row=current_row, column=4, value=item.ingredient.unit)
        ws.cell(row=current_row, column=5, value=float(item.missing_qty))
        ws.cell(row=current_row, column=6, value="-") # 주문비고
        ws.cell(row=current_row, column=7, value="-") # 원산지
        # H to M are blank
        for col_idx in range(1, 14):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

    # Fill empty rows up to 20 if needed
    while current_row < 25:
        for col_idx in range(1, 14):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = header_border
        current_row += 1

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"daily_inspection_{target_date}_{filename_suffix}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
